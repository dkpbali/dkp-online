"""
xlsx_reader.py  –  Baca data dari Kendaraan_Sekretariat.xlsx
"""
from pathlib import Path
from datetime import datetime
import openpyxl

XLSX = Path(__file__).parent.parent / "Kendaraan_Sekretariat.xlsx"


def _load(path=None):
    return openpyxl.load_workbook(str(path or XLSX), data_only=True)


def get_kendaraan(kode: str, path=None) -> dict:
    """Lookup 1 kendaraan dari Master_Kendaraan by kode."""
    wb = _load(path)
    ws = wb["Master_Kendaraan"]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if str(row[0]) == kode:
            return {
                "kode":         row[0],
                "jenis":        row[1],
                "nama":         row[2],
                "no_polisi":    row[3],
                "tahun":        row[4],
                "warna":        row[5],
            }
    raise ValueError(f"Kendaraan '{kode}' tidak ditemukan.")


def get_peminjaman(no_ba: str, path=None) -> dict:
    """Ambil 1 baris dari Pinjam_Kendaraan by No. BA."""
    wb   = _load(path)
    ws_p = wb["Pinjam_Kendaraan"]
    # Col: A=no_ba B=tgl_pinjam C=tgl_kembali D=kode E=jenis F=nama
    #      G=nip   H=tujuan     I=km_awal     J=km_akhir K=ptj  L=kondisi
    #      M=ttd_nama N=status
    for row in ws_p.iter_rows(min_row=4, values_only=True):
        if str(row[0]) == no_ba:
            kode = str(row[3]) if row[3] else ""
            try:
                kend = get_kendaraan(kode, path)
                nama_kend = kend["nama"]
                no_pol    = kend["no_polisi"]
            except:
                nama_kend, no_pol = kode, "-"

            tgl = row[1]
            return {
                "NO_BA":            str(row[0]),
                "TANGGAL_PINJAM":   tgl.strftime("%d/%m/%Y") if isinstance(tgl, datetime) else str(tgl or ""),
                "TANGGAL_KEMBALI":  str(row[2] or "-"),
                "NAMA_PEMINJAM":    str(row[5] or ""),
                "NIP_NIK":          str(row[6] or ""),
                "TUJUAN":           str(row[7] or ""),
                "NAMA_KENDARAAN":   nama_kend,
                "NO_POLISI":        no_pol,
                "JENIS_PEMINJAMAN": str(row[4] or ""),
                "JABATAN":          str(row[4] or "Pegawai"),
                "KONDISI_KEMBALI":  str(row[11] or "-"),
            }
    raise ValueError(f"No. BA '{no_ba}' tidak ditemukan.")


def list_peminjaman(path=None) -> list[dict]:
    """Semua baris Pinjam_Kendaraan (untuk dashboard)."""
    wb = _load(path)
    ws = wb["Pinjam_Kendaraan"]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]: continue
        rows.append({
            "no_ba":      row[0],
            "tgl_pinjam": str(row[1]) if row[1] else "",
            "tgl_kembali":str(row[2]) if row[2] else "",
            "kode":       row[3],
            "jenis":      row[4],
            "nama":       row[5],
            "tujuan":     row[7],
            "kondisi":    row[11],
            "status":     row[13],
        })
    return rows


def list_service(path=None) -> list[dict]:
    """Semua baris Service_Kendaraan (untuk dashboard)."""
    wb = _load(path)
    ws = wb["Service_Kendaraan"]
    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]: continue
        rows.append({
            "tgl":       str(row[0]) if row[0] else "",
            "kode":      row[1],
            "jenis":     row[2],
            "uraian":    row[3],
            "biaya":     row[4],
            "bengkel":   row[5],
            "jatuh_tempo": str(row[6]) if row[6] else "",
            "bukti":     row[7],
        })
    return rows
